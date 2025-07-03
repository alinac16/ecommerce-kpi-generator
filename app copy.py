# app.py (updated)
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
import os
import pandas as pd
from io import StringIO, BytesIO
import base64
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
import re 


from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
# TEMPORARY: Allow all origins for debugging CORS issues.
# IMPORTANT: DO NOT USE "*" IN PRODUCTION. Revert to explicit origins list.
CORS(app, origins="*") 

@app.route('/')
def serve_index():
    return render_template('test.html')

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"

if not GEMINI_API_KEY:
    print("Warning: GEMINI_API_KEY not found in environment variables. Please set it in a .env file.")
    # In a real production app, you might want to raise an error and prevent startup

# Helper function to find a column name in df.columns that matches a list of regex patterns
def find_column_by_pattern(df_columns, patterns):
    for pattern in patterns:
        for col in df_columns:
            if re.search(pattern, col, re.IGNORECASE):
                return col
    return None

# --- Re-Added: AI Column Mapping Suggestion Endpoint ---
@app.route('/suggest-column-mapping', methods=['POST'])
def suggest_column_mapping():
    if not GEMINI_API_KEY:
        return jsonify({"error": "API Key not configured on server."}), 500

    try:
        data = request.get_json()
        csv_headers = data.get('headers')

        if not csv_headers:
            return jsonify({"error": "Missing 'headers' in request."}), 400

        headers_str = ", ".join(csv_headers)
        
        # Define common conceptual column types for e-commerce
        conceptual_columns = [
            "Order ID", "Product Name", "Quantity", "Price", "Customer ID", 
            "Order Date", "Category"
        ]
        conceptual_columns_str = ", ".join(conceptual_columns)

        prompt = f"""You are an intelligent data mapping assistant for e-commerce data.
Given a list of actual CSV headers, map each relevant header to one of the following standard conceptual e-commerce column types: {conceptual_columns_str}.
Only map headers that clearly correspond to these conceptual types. If a header doesn't fit, do not include it in the mapping.
For each mapping, output a JSON object with the original CSV header as the key and the conceptual column type as the value.

Example CSV Headers: order_id, item_name, units_sold, item_price, client_id, transaction_date, product_category

Example Output (JSON):
```json
{{
  "order_id": "Order ID",
  "item_name": "Product Name",
  "units_sold": "Quantity",
  "item_price": "Price",
  "client_id": "Customer ID",
  "transaction_date": "Order Date",
  "product_category": "Category"
}}
```

Current CSV Headers: {headers_str}
Suggested Mapping (JSON):
"""
        chat_history = []
        chat_history.append({ "role": "user", "parts": [{ "text": prompt }] })

        gemini_payload = {
            "contents": chat_history,
            "generationConfig": {
                "temperature": 0.1, # Keep temperature low for consistent JSON output
            },
        }

        response = requests.post(
            f"{GEMINI_API_URL}?key={GEMINI_API_KEY}",
            headers={'Content-Type': 'application/json'},
            json=gemini_payload
        )
        response.raise_for_status()
        gemini_response = response.json()

        suggested_mapping_raw_text = gemini_response['candidates'][0]['content']['parts'][0]['text'].strip()

        # Robust JSON parsing (removing markdown if present)
        if suggested_mapping_raw_text.startswith('```json'):
            suggested_mapping_raw_text = suggested_mapping_raw_text[len('```json'):]
            if suggested_mapping_raw_text.endswith('```'):
                suggested_mapping_raw_text = suggested_mapping_raw_text[:-len('```')]
        
        suggested_mapping_raw_text = suggested_mapping_raw_text.strip()
        
        try:
            suggested_mapping = json.loads(suggested_mapping_raw_text)
            if not isinstance(suggested_mapping, dict):
                raise ValueError("AI returned unexpected JSON type, expected a dictionary.")

        except json.JSONDecodeError as e:
            print(f"JSON Decode Error in /suggest-column-mapping: {e}")
            print(f"Raw AI Response: {suggested_mapping_raw_text}")
            return jsonify({"error": "AI mapping response was not valid JSON. Please try again."}), 500
        except ValueError as e:
            print(f"AI mapping response validation error: {e}")
            return jsonify({"error": f"AI returned unexpected mapping structure: {e}. Please try again."}), 500

        return jsonify({"mapping": suggested_mapping}), 200

    except requests.exceptions.RequestException as e:
        print(f"Error calling Gemini API for column mapping suggestions: {e}")
        return jsonify({"error": f"Failed to get column mapping from AI: {e}"}), 500
    except Exception as e:
        print(f"An unexpected error occurred in /suggest-column-mapping: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Internal server error: {e}"}), 500

# --- MODIFIED: AI KPI Suggestion Endpoint (now uses conceptual headers) ---
@app.route('/suggest-kpis', methods=['POST'])
def suggest_kpis():
    if not GEMINI_API_KEY:
        return jsonify({"error": "API Key not configured on server."}), 500

    try:
        data = request.get_json()
        # Expecting the confirmed conceptual column names from the frontend
        conceptual_headers = data.get('conceptual_headers') 

        if not conceptual_headers:
            return jsonify({"error": "Missing 'conceptual_headers' in request."}), 400

        conceptual_headers_str = ", ".join(conceptual_headers)
        # MODIFIED PROMPT: Emphasize simple equations
        prompt = f"""You are an expert e-commerce data analyst assistant.
Given the following conceptual e-commerce column headers that are available in the dataset: {conceptual_headers_str},
suggest as many common and useful Key Performance Indicators (KPIs) as possible that can be calculated.
For each KPI, ensure it is calculable using ONLY the provided conceptual headers.
Provide the KPI name, a **simple and easy-to-read equation focusing only on column names and basic math operations (e.g., 'SUM(Price)', 'Price * Quantity', 'COUNT(DISTINCT Customer ID)'). Do NOT include complex functions like MONTH(), SUBSTRING(), or detailed syntax**, a brief English description of what the KPI shows and what a good number generally looks like for an an e-commerce business, and a brief Thai description.

Format your response as a JSON array of objects. Each object must have the following keys:
"name": (string, e.g., "Total Sales")
"equation": (string, e.g., "SUM(Price)")
"english_description": (string, e.g., "Total revenue generated from all sales. A higher number indicates more sales activity and revenue.")
"thai_description": (string, e.g., "รายได้รวมที่มาจากการขายทั้งหมด ตัวเลขที่สูงขึ้นแสดงถึงกิจกรรมการขายและรายได้ที่มากขึ้น")

Example Conceptual Headers: Order ID, Product Name, Quantity, Price, Customer ID, Order Date, Category

Example Output (JSON):
```json
[
  {{
    "name": "Total Sales",
    "equation": "SUM(Price)",
    "english_description": "Total revenue generated from all sales. A higher number indicates more sales activity and revenue.",
    "thai_description": "รายได้รวมที่มาจากการขายทั้งหมด ตัวเลขที่สูงขึ้นแสดงถึงกิจกรรมการขายและรายได้ที่มากขึ้น"
  }},
  {{
    "name": "Average Order Value (AOV)",
    "equation": "SUM(Price) / COUNT(DISTINCT 'Order ID')",
    "english_description": "The average amount spent per order. A higher AOV means customers are spending more each time they purchase.",
    "thai_description": "ยอดใช้จ่ายเฉลี่ยต่อคำสั่งซื้อ AOV ที่สูงขึ้นหมายความว่าลูกค้าใช้จ่ายมากขึ้นในการซื้อแต่ละครั้ง"
  }},
  {{
    "name": "Number of Unique Customers",
    "equation": "COUNT(DISTINCT 'Customer ID')",
    "english_description": "The total count of unique customers who made purchases. A growing number suggests successful customer acquisition.",
    "thai_description": "จำนวนลูกค้ารวมที่ไม่ซ้ำกันที่ทำการซื้อ ตัวเลขที่เพิ่มขึ้นแสดงถึงการหาลูกค้าใหม่ที่ประสบความสำเร็จ"
  }},
  {{
    "name": "Top Selling Products",
    "equation": "SUM(Price) GROUP BY 'Product Name' ORDER BY SUM(Price) DESC (Top 10)",
    "english_description": "A list of the products that generate the most revenue. Helps identify popular items and optimize inventory.",
    "thai_description": "รายการสินค้าที่สร้างรายได้สูงสุด ช่วยระบุสินค้าที่ได้รับความนิยมและปรับปรุงการจัดการสินค้าคงคลัง"
  }},
  {{
    "name": "Monthly Sales Trend",
    "equation": "SUM(Price) GROUP BY 'Order Date'",
    "english_description": "Sales performance over time, typically aggregated by month. Useful for identifying seasonal trends and growth patterns.",
    "thai_description": "ประสิทธิภาพการขายเมื่อเวลาผ่านไป โดยปกติจะรวมเป็นรายเดือน มีประโยชน์ในการระบุแนวโน้มตามฤดูกาลและรูปแบบการเติบโต"
  }}
]
```

Current Conceptual Headers: {conceptual_headers_str}
Suggested KPIs (JSON Array):
"""
        chat_history = []
        chat_history.append({ "role": "user", "parts": [{ "text": prompt }] })

        gemini_payload = {
            "contents": chat_history,
            "generationConfig": {
                "temperature": 0.1, # Keep temperature low for more consistent JSON/structured output
            },
        }

        response = requests.post(
            f"{GEMINI_API_URL}?key={GEMINI_API_KEY}",
            headers={'Content-Type': 'application/json'},
            json=gemini_payload
        )
        response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
        gemini_response = response.json()

        suggested_kpis_raw_text = gemini_response['candidates'][0]['content']['parts'][0]['text'].strip()

        # Attempt to parse JSON. Gemini might wrap it in markdown.
        if suggested_kpis_raw_text.startswith('```json'):
            suggested_kpis_raw_text = suggested_kpis_raw_text[len('```json'):]
            if suggested_kpis_raw_text.endswith('```'):
                suggested_kpis_raw_text = suggested_kpis_raw_text[:-len('```')]
        
        suggested_kpis_raw_text = suggested_kpis_raw_text.strip()
        
        try:
            suggested_kpis_list = json.loads(suggested_kpis_raw_text)
            # Basic validation to ensure it's a list of dicts with expected keys
            if not isinstance(suggested_kpis_list, list) or \
               not all(isinstance(item, dict) and
                       all(key in item for key in ["name", "equation", "english_description", "thai_description"])
                       for item in suggested_kpis_list):
                raise ValueError("AI returned unexpected JSON structure or missing keys.")

        except json.JSONDecodeError as e:
            print(f"JSON Decode Error: {e}")
            print(f"Raw AI Response: {suggested_kpis_raw_text}")
            return jsonify({"error": "AI response was not valid JSON. Please try again or refine headers."}), 500
        except ValueError as e:
            print(f"AI response validation error: {e}")
            return jsonify({"error": f"AI returned unexpected KPI structure: {e}. Please try again."}), 500


        return jsonify({"kpis": suggested_kpis_list}), 200

    except requests.exceptions.RequestException as e:
        print(f"Error calling Gemini API for KPI suggestions: {e}")
        return jsonify({"error": f"Failed to get KPI suggestions from AI: {e}"}), 500
    except Exception as e:
        print(f"An unexpected error occurred in /suggest-kpis: {e}")
        import traceback
        traceback.print_exc() # For debugging
        return jsonify({"error": f"Internal server error: {e}"}), 500


# --- Helper function for AI-driven data type cleaning ---
def clean_dataframe_types_with_ai_suggestions(df, gemini_api_key, gemini_api_url):
    """
    Uses Gemini AI to suggest data types for DataFrame columns and then applies them.
    """
    cleaned_df = df.copy()
    headers = df.columns.tolist()

    # Prepare prompt for Gemini AI to suggest data types
    # Asking for JSON output is crucial for reliable parsing
    prompt = f"""You are an expert data type assistant. Given the following CSV headers, suggest the most appropriate pandas data type for each column. Provide the response as a JSON object where keys are the column names and values are the suggested pandas data types (e.g., 'int64', 'float64', 'object', 'datetime64[ns]', 'bool'). If a column could be multiple types, suggest the most common and useful one based on e-commerce context.

Example Headers: ["Order ID", "Product Name", "Quantity", "Price", "Customer ID", "Order Date", "Category", "City"]

Example Output:
{{"Order ID": "object", "Product Name": "object", "Quantity": "int64", "Price": "float64", "Customer ID": "object", "Order Date": "datetime64[ns]", "Category": "object", "City": "object"}}

Current Headers: {json.dumps(headers)}
Output:
"""
    chat_history = []
    chat_history.append({"role": "user", "parts": [{"text": prompt}]})

    gemini_payload = {
        "contents": chat_history,
        "generationConfig": {
            "temperature": 0.2, # Low temperature for consistent, factual output
            # No responseSchema here for text model, we expect JSON string in text output
        },
    }

    print(f"Calling Gemini for type suggestions for headers: {headers}")

    try:
        response = requests.post(
            f"{gemini_api_url}?key={gemini_api_key}",
            headers={'Content-Type': 'application/json'},
            json=gemini_payload
        )
        response.raise_for_status() # Raise an HTTPError for bad responses (4xx or 5xx)
        gemini_response = response.json()

        suggested_types_raw = gemini_response['candidates'][0]['content']['parts'][0]['text'].strip()
        print(f"Gemini raw response for types: {suggested_types_raw}")

        # Attempt to parse the AI's JSON string response
        dtype_suggestions = json.loads(suggested_types_raw)
        print(f"Parsed AI type suggestions: {dtype_suggestions}")

    except (requests.exceptions.RequestException, KeyError, json.JSONDecodeError) as e:
        print(f"Error getting or parsing AI type suggestions: {e}")
        print("Falling back to pandas default type inference for cleaning.")
        return df.infer_objects() # Fallback to pandas' general inference if AI fails

    # Apply AI-suggested types
    for col, dtype_str in dtype_suggestions.items():
        if col not in cleaned_df.columns:
            continue # Skip if AI suggested type for a non-existent column

        try:
            # Handle common pandas data types explicitly for robustness
            if 'int' in dtype_str.lower():
                # Use 'Int64' (capital I) for nullable integer type
                cleaned_df[col] = pd.to_numeric(cleaned_df[col], errors='coerce').astype('Int64')
            elif 'float' in dtype_str.lower():
                cleaned_df[col] = pd.to_numeric(cleaned_df[col], errors='coerce')
            elif 'datetime' in dtype_str.lower():
                cleaned_df[col] = pd.to_datetime(cleaned_df[col], errors='coerce', infer_datetime_format=True)
            elif 'bool' in dtype_str.lower():
                # Convert common string representations to boolean
                cleaned_df[col] = cleaned_df[col].astype(str).str.lower().map({
                    'true': True, 'false': False, '1': True, '0': False,
                    't': True, 'f': False, 'yes': True, 'no': False
                }).fillna(pd.NA).astype('boolean') # Use nullable boolean
            elif 'object' in dtype_str.lower() or 'string' in dtype_str.lower():
                cleaned_df[col] = cleaned_df[col].astype(str) # Ensure it's string object
            else:
                # Attempt direct conversion for any other specified pandas dtype
                cleaned_df[col] = cleaned_df[col].astype(dtype_str, errors='ignore') # 'ignore' leaves as is if conversion fails
        except Exception as e:
            print(f"Failed to convert column '{col}' to '{dtype_str}' as suggested by AI: {e}. Keeping original type.")
            # If AI-suggested conversion fails, let pandas infer or keep original
            cleaned_df[col] = cleaned_df[col].infer_objects() # Try to infer best object type


    # Final pass to infer any types that AI might have missed or couldn't determine
    # This also handles columns not suggested by AI
    cleaned_df = cleaned_df.infer_objects()
    return cleaned_df


# --- Reusable DataFrame Processing Function ---
def _process_dataframe_for_kpis(csv_string, confirmed_column_mapping, gemini_api_key, gemini_api_url): # Added API key/URL
    """
    Loads CSV into DataFrame, standardizes columns, and performs data type conversions.
    Returns the processed DataFrame.
    """
    df = pd.read_csv(StringIO(csv_string))
    df.columns = df.columns.str.strip()

    # Debugging: Print DataFrame info before AI-driven conversion
    print("DataFrame columns and dtypes BEFORE AI-driven conversion:")
    print(df.info())

    # --- NEW: AI-driven data type cleaning ---
    df_cleaned = clean_dataframe_types_with_ai_suggestions(df, gemini_api_key, gemini_api_url)
    # --- END NEW ---

    # --- Post-AI cleaning, manual dropna for critical KPI columns based on mapping ---
    # These are needed for KPI calculations to avoid NaN issues
    price_col_actual = confirmed_column_mapping.get('Price')
    quantity_col_actual = confirmed_column_mapping.get('Quantity')
    order_date_col_actual = confirmed_column_mapping.get('Order Date')

    # Drop rows where critical columns became NaN after conversion
    # Only drop if the column actually exists in the dataframe after cleaning
    if price_col_actual and price_col_actual in df_cleaned.columns and pd.api.types.is_numeric_dtype(df_cleaned[price_col_actual]):
        df_cleaned.dropna(subset=[price_col_actual], inplace=True)
    if quantity_col_actual and quantity_col_actual in df_cleaned.columns and pd.api.types.is_numeric_dtype(df_cleaned[quantity_col_actual]):
        df_cleaned.dropna(subset=[quantity_col_actual], inplace=True)
    if order_date_col_actual and order_date_col_actual in df_cleaned.columns and pd.api.types.is_datetime64_any_dtype(df_cleaned[order_date_col_actual]):
        df_cleaned.dropna(subset=[order_date_col_actual], inplace=True)


    # Debugging: Print DataFrame info AFTER all processing
    print("DataFrame columns and dtypes AFTER all processing (AI + manual dropna):")
    print(df_cleaned.info())
    print("DataFrame head after all processing:")
    print(df_cleaned.head())

    return df_cleaned


# --- NEW Endpoint: Clean Data and Download Raw Excel ---
@app.route('/clean-data', methods=['POST'])
def clean_data():
    try:
        data = request.get_json()
        base64_csv_data = data.get('csv_data')
        confirmed_column_mapping = data.get('confirmed_column_mapping')

        if not base64_csv_data or not confirmed_column_mapping:
            return jsonify({"error": "Missing CSV data or column mapping."}), 400
        
        csv_bytes = base64.b64decode(base64_csv_data)
        csv_string = csv_bytes.decode('utf-8')

        # Pass API key and URL to the processing function
        df_cleaned = _process_dataframe_for_kpis(csv_string, confirmed_column_mapping, GEMINI_API_KEY, GEMINI_API_URL)

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_cleaned.to_excel(writer, sheet_name='Cleaned Raw Data', index=False)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='cleaned_raw_data.xlsx')

    except Exception as e:
        print(f"An error occurred during cleaned data generation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate cleaned data Excel file: {e}"}), 500


# --- MODIFIED: KPI Calculation and Excel Generation Endpoint (uses confirmed_column_mapping) ---
@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    try:
        data = request.get_json()
        base64_csv_data = data.get('csv_data')
        selected_kpis_meta = data.get('selected_kpis') # These are full KPI objects
        confirmed_column_mapping = data.get('confirmed_column_mapping') # User-confirmed mapping

        if not base64_csv_data or not selected_kpis_meta or not confirmed_column_mapping:
            return jsonify({"error": "Missing CSV data, selected KPIs, or column mapping."}), 400

        # Debugging: Print the received mapping
        print(f"Received confirmed_column_mapping for Excel generation: {confirmed_column_mapping}")

        csv_bytes = base64.b64decode(base64_csv_data)
        csv_string = csv_bytes.decode('utf-8')

        # Use the reusable function to get the cleaned DataFrame
        # Pass API key and URL here as well
        df = _process_dataframe_for_kpis(csv_string, confirmed_column_mapping, GEMINI_API_KEY, GEMINI_API_URL)


        # Create an Excel writer object in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Raw Data
            df.to_excel(writer, sheet_name='Raw Data', index=False)

            # Sheet 2: Calculated KPIs
            kpi_ws = writer.book.create_sheet("Calculated KPIs")
            kpi_ws.column_dimensions['A'].width = 20
            kpi_ws.column_dimensions['B'].width = 50
            kpi_ws.column_dimensions['C'].width = 50

            kpi_calculated_results = {} # Stores calculated values (scalars or dataframes)

            # Perform calculations using the confirmed_column_mapping to access df columns
            for kpi_meta in selected_kpis_meta:
                kpi_name = kpi_meta['name']

                # Retrieve actual column names from the confirmed mapping
                # These columns should now be in the correct dtype or NaN after the centralized conversion
                price_col_actual = confirmed_column_mapping.get('Price')
                quantity_col_actual = confirmed_column_mapping.get('Quantity')
                order_id_col_actual = confirmed_column_mapping.get('Order ID')
                customer_id_col_actual = confirmed_column_mapping.get('Customer ID')
                product_name_col_actual = confirmed_column_mapping.get('Product Name')
                category_col_actual = confirmed_column_mapping.get('Category')
                order_date_col_actual = confirmed_column_mapping.get('Order Date')


                if kpi_name == 'Total Sales':
                    # Check if 'Price' column exists and is numeric after conversion
                    if price_col_actual and price_col_actual in df.columns and pd.api.types.is_numeric_dtype(df[price_col_actual]) and not df[price_col_actual].empty:
                        kpi_calculated_results['Total Sales'] = df[price_col_actual].sum()
                    # Fallback for 'Quantity' * 'Price' (e.g., if 'Price' means unit price)
                    elif quantity_col_actual and price_col_actual and \
                         quantity_col_actual in df.columns and price_col_actual in df.columns and \
                         pd.api.types.is_numeric_dtype(df[quantity_col_actual]) and pd.api.types.is_numeric_dtype(df[price_col_actual]) and \
                         not df[quantity_col_actual].empty and not df[price_col_actual].empty:
                        kpi_calculated_results['Total Sales'] = (df[quantity_col_actual] * df[price_col_actual]).sum()
                    else:
                        kpi_calculated_results['Total Sales'] = None

                if kpi_name == 'Total Quantity Sold':
                    if quantity_col_actual and quantity_col_actual in df.columns and pd.api.types.is_numeric_dtype(df[quantity_col_actual]) and not df[quantity_col_actual].empty:
                        kpi_calculated_results['Total Quantity Sold'] = df[quantity_col_actual].sum()
                    else:
                        kpi_calculated_results['Total Quantity Sold'] = None

                if kpi_name == 'Average Order Value (AOV)':
                    if price_col_actual and order_id_col_actual and \
                       price_col_actual in df.columns and order_id_col_actual in df.columns and \
                       pd.api.types.is_numeric_dtype(df[price_col_actual]) and not df[price_col_actual].empty:
                        sales_per_order = df.groupby(order_id_col_actual)[price_col_actual].sum()
                        if not sales_per_order.empty:
                            kpi_calculated_results['Average Order Value (AOV)'] = sales_per_order.mean()
                        else:
                            kpi_calculated_results['Average Order Value (AOV)'] = None
                    else:
                        kpi_calculated_results['Average Order Value (AOV)'] = None

                if kpi_name == 'Number of Unique Customers':
                    if customer_id_col_actual and customer_id_col_actual in df.columns and not df[customer_id_col_actual].empty:
                        kpi_calculated_results['Number of Unique Customers'] = df[customer_id_col_actual].nunique()
                    else:
                        kpi_calculated_results['Number of Unique Customers'] = None

                if kpi_name == 'Top Selling Products':
                    if product_name_col_actual and price_col_actual and \
                       product_name_col_actual in df.columns and price_col_actual in df.columns and \
                       pd.api.types.is_numeric_dtype(df[price_col_actual]) and not df[price_col_actual].empty:
                        top_products = df.groupby(product_name_col_actual)[price_col_actual].sum().nlargest(10).sort_values(ascending=False)
                        if not top_products.empty:
                            kpi_calculated_results['Top Selling Products'] = top_products.rename('Total Sales').to_frame()
                        else:
                            kpi_calculated_results['Top Selling Products'] = None
                    else:
                        kpi_calculated_results['Top Selling Products'] = None

                if kpi_name == 'Sales by Product Category':
                    if category_col_actual and price_col_actual and \
                       category_col_actual in df.columns and price_col_actual in df.columns and \
                       pd.api.types.is_numeric_dtype(df[price_col_actual]) and not df[price_col_actual].empty:
                        sales_by_category = df.groupby(category_col_actual)[price_col_actual].sum().sort_values(ascending=False)
                        if not sales_by_category.empty:
                            kpi_calculated_results['Sales by Product Category'] = sales_by_category.rename('Total Sales').to_frame()
                        else:
                            kpi_calculated_results['Sales by Product Category'] = None
                    else:
                        kpi_calculated_results['Sales by Product Category'] = None

                if kpi_name == 'Monthly Sales Trend':
                    if order_date_col_actual and price_col_actual and \
                       order_date_col_actual in df.columns and price_col_actual in df.columns and \
                       pd.api.types.is_datetime64_any_dtype(df[order_date_col_actual]) and pd.api.types.is_numeric_dtype(df[price_col_actual]) and \
                       not df[order_date_col_actual].empty and not df[price_col_actual].empty:
                        
                        df_monthly = df.set_index(order_date_col_actual)[price_col_actual].resample('M').sum().to_frame(name='Total Sales')
                        if not df_monthly.empty:
                            df_monthly.index = df_monthly.index.strftime('%Y-%m')
                            kpi_calculated_results['Monthly Sales Trend'] = df_monthly
                        else:
                            kpi_calculated_results['Monthly Sales Trend'] = None
                    else:
                        kpi_calculated_results['Monthly Sales Trend'] = None

            # --- Writing KPIs to Excel Sheet and Adding Charts ---
            current_row = 1 # Start writing from row 1

            for kpi_meta in selected_kpis_meta:
                kpi_name = kpi_meta['name']
                kpi_equation = kpi_meta['equation']
                kpi_eng_desc = kpi_meta['english_description']
                kpi_thai_desc = kpi_meta['thai_description']
                
                calculated_value = kpi_calculated_results.get(kpi_name)

                # Write KPI Name
                kpi_ws.cell(row=current_row, column=1, value=kpi_name).font = Font(bold=True, size=14, color="000000")
                current_row += 1

                if calculated_value is None:
                    kpi_ws.cell(row=current_row, column=1, value="Status:").font = Font(bold=True)
                    kpi_ws.cell(row=current_row, column=2, value="Not enough data or required columns to calculate.").font = Font(color="FF0000")
                    current_row += 2
                    kpi_ws.cell(row=current_row, column=1).value = "" # Add blank row for spacing
                    current_row += 1
                    continue # Skip to next KPI if not not calculated

                # Write Equation
                kpi_ws.cell(row=current_row, column=1, value="Equation:").font = Font(bold=True)
                kpi_ws.cell(row=current_row, column=2, value=kpi_equation).font = Font(italic=True, color="0000FF")
                current_row += 1

                # Write English Description
                kpi_ws.cell(row=current_row, column=1, value="English:").font = Font(bold=True)
                kpi_ws.cell(row=current_row, column=2, value=kpi_eng_desc)
                kpi_ws.row_dimensions[current_row].height = 40
                kpi_ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1

                # Write Thai Description
                kpi_ws.cell(row=current_row, column=1, value="ภาษาไทย:").font = Font(bold=True)
                kpi_ws.cell(row=current_row, column=2, value=kpi_thai_desc)
                kpi_ws.row_dimensions[current_row].height = 40
                kpi_ws.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
                current_row += 1

                # Write the calculated value or DataFrame
                if isinstance(calculated_value, (int, float)):
                    kpi_ws.cell(row=current_row, column=1, value="Calculated Value:").font = Font(bold=True)
                    kpi_ws.cell(row=current_row, column=2, value=calculated_value).font = Font(size=12)
                    current_row += 2
                elif isinstance(calculated_value, pd.DataFrame):
                    # Write DataFrame header and data
                    df_start_row_for_chart = current_row
                    for r_idx, row_data in enumerate(dataframe_to_rows(calculated_value, index=True, header=True)):
                        for c_idx, cell_value in enumerate(row_data):
                            kpi_ws.cell(row=current_row + r_idx, column=c_idx + 1, value=cell_value)
                    
                    df_end_row_for_chart = current_row + len(calculated_value) # Last row of data (including header)
                    df_end_col_for_chart = len(calculated_value.columns) + 1 # Last column of data (including index)

                    current_row += len(calculated_value) + 2 # Move past dataframe rows + header + 1 blank row


                    # --- Add Chart for DataFrame KPIs ---
                    if kpi_name in ["Top Selling Products", "Sales by Product Category"]:
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        chart.title = kpi_name
                        chart.y_axis.title = calculated_value.columns[0] if len(calculated_value.columns) > 0 else "Value"
                        chart.x_axis.title = calculated_value.index.name if calculated_value.index.name else "Category/Product"

                        data = Reference(kpi_ws, min_col=df_end_col_for_chart, min_row=df_start_row_for_chart + 1, max_col=df_end_col_for_chart, max_row=df_end_row_for_chart)
                        categories = Reference(kpi_ws, min_col=1, min_row=df_start_row_for_chart + 1, max_col=1, max_row=df_end_row_for_chart)
                        
                        chart.add_data(data, titles_from_data=False)
                        chart.set_categories(categories)
                        
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showVal = True

                        kpi_ws.add_chart(chart, f"D{current_row}")
                        current_row += 15
                        
                    elif kpi_name == "Monthly Sales Trend":
                        chart = LineChart()
                        chart.style = 10
                        chart.title = "Monthly Sales Trend"
                        chart.y_axis.title = calculated_value.columns[0] if len(calculated_value.columns) > 0 else "Value"
                        chart.x_axis.title = "Month"

                        data = Reference(kpi_ws, min_col=df_end_col_for_chart, min_row=df_start_row_for_chart + 1, max_col=df_end_col_for_chart, max_row=df_end_row_for_chart)
                        categories = Reference(kpi_ws, min_col=1, min_row=df_start_row_for_chart + 1, max_col=1, max_row=df_end_row_for_chart)
                        
                        chart.add_data(data, titles_from_data=False)
                        chart.set_categories(categories)
                        
                        chart.dataLabels = DataLabelList()
                        chart.dataLabels.showVal = True

                        kpi_ws.add_chart(chart, f"D{current_row}")
                        current_row += 15

                current_row += 2 # Add some space before the next KPI section

            # Auto-size columns for better readability (only for columns that had content written)
            for i, col in enumerate(kpi_ws.columns):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            cell_text = str(cell.value)
                            if len(cell_text) > 40:
                                max_length = 40
                                break
                            max_length = max(max_length, len(cell_text))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                kpi_ws.column_dimensions[column].width = max(adjusted_width, 10)


            # Ensure 'Calculated KPIs' is the second sheet if it's new
            if 'Calculated KPIs' in writer.book.sheetnames:
                kpi_sheet_index = writer.book.sheetnames.index('Calculated KPIs')
                writer.book.move_sheet(writer.book.worksheets[kpi_sheet_index], offset=1 - kpi_sheet_index)
            
            # Remove default empty sheet if it exists (usually "Sheet" or "Sheet1")
            if "Sheet" in writer.book.sheetnames:
                writer.book.remove(writer.book["Sheet"])
            if "Sheet1" in writer.book.sheetnames:
                writer.book.remove(writer.book["Sheet1"])

        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='ecommerce_kpis.xlsx')

    except Exception as e:
        print(f"An error occurred during Excel generation: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate Excel file: {e}"}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5000)
